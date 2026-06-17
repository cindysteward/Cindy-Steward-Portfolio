"""
Local web UI for FACET-CV.

This is a browser-based form to configure a study-prompter analysis session,
upload input files, monitor pipeline progress, and browse output results.
Requires Flask.  Run from the project root:

    python launch_ui.py [--port 5050]

Then open http://localhost:5050 in your browser.
"""
import matplotlib
matplotlib.use('Agg')

import sys
import argparse
import threading
import queue
import json
import time
import socket
import ssl
import uuid
import webbrowser
from datetime import datetime
from pathlib import Path
from typing import Optional
import urllib.request as _ureq

from flask import Flask, render_template, request, jsonify, Response, send_file

_this_dir = Path(__file__).resolve().parent
_candidate_child = _this_dir / "master_project"
_candidate_sibling = _this_dir.parent / "master_project"
if _candidate_child.exists():
    PIPELINE_ROOT = _candidate_child
elif _candidate_sibling.exists():
    PIPELINE_ROOT = _candidate_sibling
else:
    PIPELINE_ROOT = _this_dir
PROJECT_ROOT = PIPELINE_ROOT
sys.path.insert(0, str(PIPELINE_ROOT))

if not (PIPELINE_ROOT / "src").exists() and not (PIPELINE_ROOT / "pipeline").exists():
    print(
        f"[ERROR] Cannot find pipeline source package. Expected 'src/' or 'pipeline/' under: {PIPELINE_ROOT}"
    )
    print("        Make sure launch_ui.py is in the master_project directory.")

app = Flask(__name__, template_folder=str(PIPELINE_ROOT / "templates"))
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024
UPLOAD_FOLDER = PIPELINE_ROOT / "ui_uploads"
UPLOAD_FOLDER.mkdir(exist_ok=True)

_jobs: dict = {}
_jobs_lock = threading.Lock()
_MAX_JOBS = 20

_model_ready = threading.Event()
_model_error: Optional[str] = None


_TIMESTAMPS_CSV_COLS = [
    "participant_id", "profile", "session_date", "section", "task_number",
    "sequence", "task_type", "label", "profile_label", "expression",
    "event", "time_from_start_s", "wall_time", "recording_perf_ms", "sequence_rep",
]


def _parse_player_time(s: str) -> Optional[float]:
    """Convert a video-player timestamp string to seconds.

    Accepts M:SS, M:SS.f, H:MM:SS, H:MM:SS.f (colon-separated).
    Returns None if the string cannot be parsed.
    """
    s = s.strip()
    if not s:
        return None
    parts = s.split(":")
    try:
        if len(parts) == 3:
            return float(parts[0]) * 3600 + float(parts[1]) * 60 + float(parts[2])
        elif len(parts) == 2:
            return float(parts[0]) * 60 + float(parts[1])
        else:
            return float(parts[0])
    except (ValueError, IndexError):
        return None


def _convert_timestamps_xlsx_to_csv(xlsx_path: Path) -> Path:
    """Convert a timestamp skeleton Excel file to a pipeline-compatible CSV.

    Finds the header row by scanning for 'participant_id', extracts only the
    standard CSV columns (A-O).  Resolves time_from_start_s with this priority:
      1. Column L  (time_from_start_s — direct seconds)
      2. Column T  (→ seconds from player time, Excel formula; or Python fallback)
      3. Column Q  (→ seconds from frame number / FPS)
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required to read .xlsx files: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or "").strip().lower() == "participant_id":
                header_row = cell.row
                break
        if header_row:
            break

    if header_row is None:
        raise ValueError("Cannot find 'participant_id' header in the Excel file.")

    headers = [str(ws.cell(row=header_row, column=c).value or "").strip()
               for c in range(1, ws.max_column + 1)]

    def col_idx(name):
        try:
            return headers.index(name)
        except ValueError:
            pass
        name_lower = name.lower()
        for i, h in enumerate(headers):
            if name_lower in h.lower():
                return i
        return None

    fps_col          = col_idx("video_fps")
    frame_col        = col_idx("frame_number")
    frame_secs_col   = col_idx("= frame / fps")
    player_time_col  = col_idx("video_player_time")
    player_secs_col  = col_idx("from player time")

    csv_path = xlsx_path.with_suffix(".csv")
    import csv as _csv
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        writer = _csv.writer(fh)
        writer.writerow(_TIMESTAMPS_CSV_COLS)
        for r in range(header_row + 1, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c + 1).value for c in range(len(headers))]
            if all(v is None or str(v).strip() == "" for v in row_vals):
                continue
            out = []
            for col_name in _TIMESTAMPS_CSV_COLS:
                idx = col_idx(col_name)
                val = row_vals[idx] if idx is not None else ""
                if col_name == "time_from_start_s" and (val is None or str(val).strip() == ""):
                    if player_secs_col is not None:
                        val = row_vals[player_secs_col]
                        import datetime as _dt
                        if isinstance(val, _dt.time):
                            val = round(val.hour * 60 + val.minute + val.second / 60, 3)
                        elif isinstance(val, (int, float)) and 0 < float(val) <= 1:
                            t = _dt.datetime(1899, 12, 30) + _dt.timedelta(days=float(val))
                            val = round(t.hour * 60 + t.minute + t.second / 60, 3)
                    if (val is None or str(val).strip() == "") and player_time_col is not None:
                        raw = row_vals[player_time_col]
                        import datetime as _dt
                        if isinstance(raw, _dt.time):
                            val = round(raw.hour * 60 + raw.minute + raw.second / 60, 3)
                        elif isinstance(raw, (int, float)) and 0 < float(raw) <= 1:
                            t = _dt.datetime(1899, 12, 30) + _dt.timedelta(days=float(raw))
                            val = round(t.hour * 60 + t.minute + t.second / 60, 3)
                        elif raw not in (None, ""):
                            val = _parse_player_time(str(raw))
                    if val is None or str(val).strip() == "":
                        if frame_secs_col is not None:
                            val = row_vals[frame_secs_col]
                        if (val is None or str(val).strip() == "") and frame_col is not None and fps_col is not None:
                            frame = row_vals[frame_col]
                            fps   = row_vals[fps_col]
                            if frame not in (None, "") and fps not in (None, "", 0):
                                try:
                                    val = round(float(frame) / float(fps), 3)
                                except (TypeError, ValueError):
                                    val = ""
                out.append("" if val is None else val)
            time_idx = _TIMESTAMPS_CSV_COLS.index("time_from_start_s")
            if str(out[time_idx]).strip() == "":
                continue
            writer.writerow(out)

    return csv_path


def _extract_fps_from_timestamps(timestamps_path: Path) -> Optional[float]:
    """Read the video_fps value from the first data row of a timestamps CSV."""
    import csv as _csv
    try:
        with open(timestamps_path, newline="", encoding="utf-8-sig") as fh:
            reader = _csv.DictReader(fh)
            for row in reader:
                fps = row.get("video_fps", "").strip()
                if fps:
                    return float(fps)
    except Exception:
        pass
    return None


def _auto_generate_recording_meta(timestamps_path: Path, upload_dir: Path) -> Optional[Path]:
    """Generate a minimal recording_meta JSON from a timestamps CSV and save it.

    Extracts participant_id, session_date, profile, and video_fps from the CSV
    so the pipeline can use the correct frame rate and log session context.
    Returns the path to the saved JSON, or None on failure.
    """
    import csv as _csv
    try:
        with open(timestamps_path, newline="", encoding="utf-8-sig") as fh:
            reader = _csv.DictReader(fh)
            row = next(iter(reader), None)
        if row is None:
            return None

        fps_val = row.get("video_fps", "").strip()
        fps = float(fps_val) if fps_val else None

        camera_entry: dict = {"camera_index": 1, "start_offset_from_first_cam_s": 0}
        if fps and fps > 0:
            camera_entry["grantedFrameRate"] = fps

        meta = {
            "auto_generated": True,
            "participant_id":  row.get("participant_id", "").strip(),
            "session_date":    row.get("session_date", "").strip(),
            "profile":         row.get("profile", "").strip(),
            "cameras": [camera_entry],
        }
        out_path = upload_dir / "recording_meta_auto.json"
        out_path.write_text(json.dumps(meta, indent=2), encoding="utf-8")
        return out_path
    except Exception:
        return None


def _find_free_port(preferred: int, max_tries: int = 10) -> int:
    """Find a free port starting from `preferred`. Raises RuntimeError if none found."""
    for port in range(preferred, preferred + max_tries):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                s.bind(("127.0.0.1", port))
                return port
            except OSError:
                continue
    raise RuntimeError(f"No free port found in range {preferred}–{preferred + max_tries}")


def _wait_for_server(url: str, timeout: float = 15.0) -> bool:
    """Poll a /health endpoint until the server responds or timeout expires."""
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        try:
            _ureq.urlopen(url + "/health", timeout=1)
            return True
        except Exception:
            time.sleep(0.25)
    return False


def _prefetch_model() -> None:
    """Download the MediaPipe model in a background thread."""
    global _model_error
    try:
        from src.utils import ensure_model_downloaded
        ensure_model_downloaded()
        _model_ready.set()
    except Exception as exc:
        _model_error = str(exc)
        _model_ready.set()


def _run_pipeline_thread(
    job_id,
    video_paths,
    timestamps_path,
    meta_path,
    assembly_path,
    subject_id,
    session_label,
    study_mode,
    reference_session,
    export_db,
    video_mode,
):
    """Run the prompter pipeline in a background thread and push progress events."""
    job = _jobs[job_id]
    q: queue.Queue = job["queue"]
    try:
        try:
            from src.prompter_pipeline import run_prompter_session
        except ImportError as ie:
            q.put({"step": f"Error: Pipeline import failed: {ie}", "pct": -1, "traceback": None})
            return

        def progress_callback(step: str, pct: int) -> None:
            try:
                q.put_nowait({"step": step, "pct": pct})
            except queue.Full:
                pass

        summary = run_prompter_session(
            video_paths=video_paths,
            timestamps_path=timestamps_path,
            subject_id=subject_id,
            session_label=session_label,
            study_mode=study_mode,
            project_root=PIPELINE_ROOT,
            meta_path=meta_path,
            assembly_path=assembly_path,
            reference_session=reference_session,
            export_db=export_db,
            progress_callback=progress_callback,
            video_mode=video_mode,
        )
        job["summary"] = summary
        q.put({"step": "done", "pct": 100}, timeout=10)
    except Exception as exc:
        import traceback as _tb
        full_tb = _tb.format_exc()
        import sys as _sys
        print(full_tb, file=_sys.stderr, flush=True)
        short_msg = str(exc).split("\n")[0][:300]
        try:
            q.put({"step": f"Error: {short_msg}", "pct": -1, "traceback": full_tb}, timeout=10)
        except queue.Full:
            pass
    finally:
        job["running"] = False


@app.route("/", methods=["GET"])
def index():
    """Render the main UI page."""
    return render_template("index.html")


@app.route("/health", methods=["GET"])
def health():
    """Health check endpoint for server readiness probing."""
    return "ok", 200


@app.route("/api/model_status", methods=["GET"])
def model_status():
    """Return the model download status and any error message."""
    return jsonify({
        "ready": _model_ready.is_set(),
        "error": _model_error,
    })


@app.route("/run", methods=["POST"])
def run_analysis():
    """Accept uploaded files and form fields, then launch the pipeline in a background thread.

    Saves uploaded files to ui_uploads/<subject_id>_<timestamp>/, then spawns
    a pipeline thread that calls run_prompter_session with a progress_callback
    that pushes (step, pct) tuples to a per-job queue.  Returns immediately
    with {"status": "started", "job_id": <uuid>}.  Multiple sessions can run
    concurrently; each is tracked via its own job_id.
    """
    subject_id = request.form.get("subject_id", "").strip()
    session_label = request.form.get("session_label", "").strip()
    study_mode = request.form.get("study_mode", "pilot").strip()
    export_db = request.form.get("export_db") == "on"
    video_mode = request.form.get("video_mode", "none").strip()
    if video_mode not in ("none", "annotated", "landmark", "both"):
        video_mode = "none"

    reference_sessions_raw = request.form.get("reference_sessions", "").strip()
    reference_session = (
        [r.strip() for r in reference_sessions_raw.split(",") if r.strip()]
        if reference_sessions_raw
        else None
    )

    if not subject_id:
        return jsonify({"status": "error", "message": "Subject ID is required"}), 400
    if not session_label:
        return jsonify({"status": "error", "message": "Session label is required"}), 400

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    upload_dir = UPLOAD_FOLDER / f"{subject_id}_{ts}"
    upload_dir.mkdir(parents=True, exist_ok=True)

    video_path_strs = request.form.get("camera_video_paths", "").strip()
    if not video_path_strs:
        return jsonify({"status": "error", "message": "At least one camera video path is required"}), 400

    video_paths = []
    for p in video_path_strs.split("\n"):
        p = p.strip()
        if not p:
            continue
        vp = Path(p)
        if not vp.exists():
            return jsonify({"status": "error", "message": f"Video file not found: {p}"}), 400
        video_paths.append(vp)

    if len(video_paths) > 4:
        return jsonify({"status": "error", "message": "Maximum 4 camera video paths allowed"}), 400

    timestamps_file = request.files.get("timestamps_csv")
    if timestamps_file and timestamps_file.filename:
        timestamps_path = upload_dir / timestamps_file.filename
        timestamps_file.save(str(timestamps_path))
        if timestamps_path.suffix.lower() in (".xlsx", ".xls"):
            try:
                timestamps_path = _convert_timestamps_xlsx_to_csv(timestamps_path)
            except Exception as exc:
                return jsonify({"status": "error", "message": f"Could not convert Excel timestamps: {exc}"}), 400
    else:
        timestamps_path = None

    meta_path: Optional[Path] = None
    meta_file = request.files.get("recording_meta")
    if meta_file and meta_file.filename:
        meta_path = upload_dir / meta_file.filename
        meta_file.save(str(meta_path))
    elif timestamps_path is not None:
        meta_path = _auto_generate_recording_meta(timestamps_path, upload_dir)

    assembly_path: Optional[Path] = None
    assembly_file = request.files.get("assembly_csv")
    if assembly_file and assembly_file.filename:
        assembly_path = upload_dir / assembly_file.filename
        assembly_file.save(str(assembly_path))

    job_id = str(uuid.uuid4())
    job: dict = {"queue": queue.Queue(maxsize=500), "thread": None, "running": True, "summary": {}}

    with _jobs_lock:
        _jobs[job_id] = job
        if len(_jobs) > _MAX_JOBS:
            finished = [jid for jid, j in _jobs.items() if not j["running"] and jid != job_id]
            for jid in finished[: len(_jobs) - _MAX_JOBS]:
                del _jobs[jid]

    t = threading.Thread(
        target=_run_pipeline_thread,
        args=(
            job_id,
            video_paths,
            timestamps_path,
            meta_path,
            assembly_path,
            subject_id,
            session_label,
            study_mode,
            reference_session,
            export_db,
            video_mode,
        ),
        daemon=True,
    )
    job["thread"] = t
    t.start()
    return jsonify({"status": "started", "job_id": job_id})


@app.route("/progress/<job_id>", methods=["GET"])
def progress_stream(job_id: str):
    """Server-Sent Events endpoint that streams progress for a specific job.

    Yields data events as JSON strings.  Closes when a 'done' or 'Error:' step
    is received or when 15 minutes pass without a real event.
    """
    job = _jobs.get(job_id)
    if job is None:
        def _not_found():
            yield f"data: {json.dumps({'step': 'Error: unknown job_id', 'pct': -1})}\n\n"
        return Response(_not_found(), mimetype="text/event-stream")

    q: queue.Queue = job["queue"]

    def generate():
        deadline = time.time() + 3600
        last_real_emit = time.time()
        while time.time() < deadline:
            try:
                event = q.get(timeout=5)
                payload = json.dumps(event)
                yield f"data: {payload}\n\n"
                last_real_emit = time.time()
                step = event.get("step", "")
                if step == "done" or step.startswith("Error:"):
                    break
            except queue.Empty:
                t = job.get("thread")
                if not job["running"] and (t is None or not t.is_alive()):
                    yield f"data: {json.dumps({'step': 'done', 'pct': 100})}\n\n"
                    return
                if time.time() - last_real_emit > 900:
                    yield f"data: {json.dumps({'step': 'Error: pipeline timed out (no progress for 15 min)', 'pct': -1})}\n\n"
                    return
                yield f"data: {json.dumps({'step': 'working', 'pct': -1})}\n\n"

    return Response(generate(), mimetype="text/event-stream")


@app.route("/results/<job_id>", methods=["GET"])
def results(job_id: str):
    """Return the pipeline summary for a specific job."""
    job = _jobs.get(job_id)
    if job is None:
        return jsonify({"error": "unknown job_id"}), 404
    return jsonify(job["summary"])


@app.route("/api/jobs", methods=["GET"])
def list_jobs():
    """Return status of all known jobs (running or finished)."""
    with _jobs_lock:
        return jsonify({
            jid: {
                "running": j["running"],
                "has_summary": bool(j["summary"]),
            }
            for jid, j in _jobs.items()
        })


@app.route("/sessions", methods=["GET"])
def list_sessions():
    """Return recorded sessions for a given subject/mode as JSON.

    Query params: subject_id (required), study_mode (optional, default 'pilot').
    """
    subject_id = request.args.get("subject_id", "").strip()
    study_mode = request.args.get("study_mode", "pilot").strip()
    if not subject_id:
        return jsonify({"error": "subject_id is required"}), 400
    try:
        from src.io_manager import IOManager
        tmp_io = IOManager(PIPELINE_ROOT, subject_id, "_list", study_mode, list_only=True)
        sessions = tmp_io.list_sessions_with_metadata(subject_id, study_mode)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500
    return jsonify({"sessions": sessions})


@app.route("/visualizations/<path:filepath>", methods=["GET"])
def serve_visualization(filepath: str):
    """Serve visualization images from inside data/.

    filepath is relative to data/ (e.g. results/pilot/P001/session/visualizations/foo.png).
    """
    data_root = (PIPELINE_ROOT / "data").resolve()
    requested = (data_root / filepath).resolve()
    try:
        requested.relative_to(data_root)
    except ValueError:
        return jsonify({"error": "Access denied"}), 403
    if not requested.exists():
        return jsonify({"error": "File not found"}), 404
    return send_file(str(requested))


@app.route("/download/<path:filepath>", methods=["GET"])
def download_file(filepath: str):
    """Serve a file from inside PROJECT_ROOT/master_project/data/ for download.

    Restricts serving to the data/ subdirectory to prevent directory traversal.
    """
    data_root = (PIPELINE_ROOT / "data").resolve()
    requested = (data_root / filepath).resolve()
    try:
        requested.relative_to(data_root)
    except ValueError:
        return jsonify({"error": "Access denied"}), 403
    if not requested.exists():
        return jsonify({"error": "File not found"}), 404
    return send_file(str(requested), as_attachment=True)


@app.route("/api/normative_status", methods=["GET"])
def normative_status():
    """Return status of normative reference file availability."""
    data_dir = PIPELINE_ROOT / "data"
    normative_path = data_dir / "normative_reference.json"
    status = {
        "has_normative": normative_path.exists(),
        "normative_path": str(normative_path) if normative_path.exists() else None,
        "last_modified": None,
    }
    if normative_path.exists():
        status["last_modified"] = datetime.fromtimestamp(normative_path.stat().st_mtime).isoformat()
    return jsonify(status)


@app.route("/api/kinematic_profiles_status", methods=["GET"])
def kinematic_profiles_status():
    """Return status of kinematic reference profiles availability."""
    data_dir = PIPELINE_ROOT / "data"
    matches = sorted(data_dir.glob("results/*/*_kinematic_reference_profiles.json"))
    profiles_path = matches[0] if matches else None
    status = {
        "has_profiles": profiles_path is not None,
        "profiles_path": str(profiles_path) if profiles_path else None,
        "last_modified": None,
    }
    if profiles_path is not None:
        status["last_modified"] = datetime.fromtimestamp(profiles_path.stat().st_mtime).isoformat()
    return jsonify(status)


@app.route("/api/build_normative", methods=["POST"])
def build_normative():
    """Build normative reference from all available sessions."""
    try:
        from src.run_pipeline import _build_normative_command
        result = _build_normative_command(
            PIPELINE_ROOT / "data",
            PIPELINE_ROOT / "data",
        )
        return jsonify({
            "success": True,
            "message": "Normative reference built successfully.",
            "result": result,
        })
    except Exception as exc:
        return jsonify({
            "success": False,
            "error": str(exc),
        }), 500


@app.route("/api/list_sessions", methods=["GET"])
def api_list_sessions():
    """List all available sessions with their result paths."""
    sessions = []
    data_dir = PIPELINE_ROOT / "data" / "processed"
    if data_dir.exists():
        for mode_dir in data_dir.glob("*"):
            if not mode_dir.is_dir():
                continue
            for subject_dir in mode_dir.glob("*"):
                if not subject_dir.is_dir():
                    continue
                for session_dir in subject_dir.glob("*"):
                    if not session_dir.is_dir():
                        continue
                    results_dir = PIPELINE_ROOT / "data" / "results" / mode_dir.name / subject_dir.name / session_dir.name
                    sessions.append({
                        "session_id": session_dir.name,
                        "subject_id": subject_dir.name,
                        "study_mode": mode_dir.name,
                        "results_dir": str(results_dir) if results_dir.exists() else None,
                    })
    return jsonify({"sessions": sessions})


@app.route("/api/continuous_anomaly/<session_id>", methods=["GET"])
def continuous_anomaly_report(session_id: str):
    """Serve the continuous anomaly report JSON for a session."""
    data_dir = PIPELINE_ROOT / "data" / "results"
    for mode_dir in data_dir.glob("*"):
        for subject_dir in mode_dir.glob("*"):
            session_dir = subject_dir / session_id
            if session_dir.exists():
                report_path = session_dir / "continuous_anomaly_report.json"
                if report_path.exists():
                    try:
                        with open(report_path) as f:
                            report = json.load(f)
                        return jsonify(report)
                    except Exception as exc:
                        return jsonify({"error": str(exc)}), 500
    return jsonify({"error": "Session or report not found"}), 404


@app.route("/api/kinematic_pdf/<session_id>", methods=["GET"])
def kinematic_pdf(session_id: str):
    """Serve the kinematic profiles PDF for a session."""
    data_dir = PIPELINE_ROOT / "data" / "results"
    for mode_dir in data_dir.glob("*"):
        for subject_dir in mode_dir.glob("*"):
            session_dir = subject_dir / session_id
            if session_dir.exists():
                pdf_dir = session_dir / "kinematic_profiles"
                if pdf_dir.exists():
                    pdf_files = sorted(pdf_dir.glob("*.pdf"))
                    if pdf_files:
                        pdf_path = pdf_files[0]
                        try:
                            return send_file(str(pdf_path), mimetype="application/pdf")
                        except Exception as exc:
                            return jsonify({"error": str(exc)}), 500
                viz_dir = session_dir / "visualizations"
                if viz_dir.exists():
                    pdf_files = sorted(viz_dir.glob("kinematic*.pdf"))
                    if pdf_files:
                        try:
                            return send_file(str(pdf_files[0]), mimetype="application/pdf")
                        except Exception as exc:
                            return jsonify({"error": str(exc)}), 500
                return jsonify({"error": "Kinematic PDF not yet generated for this session. Run the analysis first."}), 404
    return jsonify({"error": "Session not found"}), 404


def _run_raw_export_worker(job_id: str, subject_id: str, session_label: str, study_mode: str,
                            video_paths: list, timestamps_path: Optional[Path],
                            meta_path: Optional[Path], q: queue.Queue):
    """Background worker for raw landmark extraction with progress updates."""
    import io as _io_mod
    
    try:
        def progress_callback(frame_idx, total_frames):
            """Called by processor every 30 frames."""
            if total_frames > 0:
                pct = int(100 * frame_idx / total_frames)
                q.put({"step": f"Extracting frame {frame_idx}/{total_frames}", "pct": min(pct, 99)})
        
        from src.multi_camera_processor import MultiCameraProcessor
        from src.study_prompter_reader import load_prompter_session
        from src.io_manager import IOManager
        import pandas as pd

        q.put({"step": "Loading session configuration...", "pct": 5})
        
        if timestamps_path is not None:
            session = load_prompter_session(
                timestamps_path=timestamps_path,
                meta_path=meta_path,
                assembly_path=None,
            )
            events_df = session.events_df
            recording_start_offset_s = session.recording_start_offset_s
        else:
            from datetime import datetime as _dt
            events_df = pd.DataFrame([
                {"timestamp_abs": 0.0, "event_type": "neutral",
                 "task_group": "0", "task_id": 0, "task_name": "continuous", "repetition": 1},
                {"timestamp_abs": 5.0, "event_type": "measurement",
                 "task_group": "0", "task_id": 1, "task_name": "continuous", "repetition": 1},
            ])
            recording_start_offset_s = 0.0

        q.put({"step": "Loading feature configuration...", "pct": 10})
        config_io = IOManager(PIPELINE_ROOT, subject_id, session_label, study_mode, list_only=True)
        features_config = config_io.load_config("features")
        
        q.put({"step": "Ensuring MediaPipe model is available...", "pct": 15})
        from src.utils import ensure_model_downloaded
        model_path = ensure_model_downloaded()

        q.put({"step": "Initializing video processor...", "pct": 20})
        processor = MultiCameraProcessor(video_paths, features_config, model_path)

        q.put({"step": "Extracting landmarks from frames...", "pct": 25})
        frame_data_list, _, _, _ = processor.process_all_frames(
            events_df,
            recording_start_offset_s=recording_start_offset_s,
            progress_callback=progress_callback,
        )

        q.put({"step": "Compiling results into DataFrame...", "pct": 95})
        df = pd.DataFrame(frame_data_list)
        df = df.drop(columns=[c for c in ["_landmarks_3d"] if c in df.columns], errors="ignore")

        ts_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{subject_id}_{session_label}_{ts_str}_raw_landmarks.csv"

        q.put({"step": "Saving to project directory...", "pct": 98})
        raw_subj_dir = PIPELINE_ROOT / "data" / "raw" / study_mode / subject_id
        raw_subj_dir.mkdir(parents=True, exist_ok=True)
        saved_path = raw_subj_dir / filename
        df.to_csv(str(saved_path), index=False)

        q.put({"step": "Done!", "pct": 100, "filename": filename, "job_id": job_id})
        
    except Exception as exc:
        q.put({"step": f"Error: {str(exc)}", "pct": -1})
    finally:
        with _jobs_lock:
            if job_id in _jobs:
                _jobs[job_id]["running"] = False


@app.route("/api/run_raw_export", methods=["POST"])
def run_raw_export():
    """Extract raw MediaPipe landmarks from video(s) in a background thread.

    Returns immediately with a job_id so the client can stream progress via SSE.
    Accepts the same video path form field as /run but skips full pipeline
    analysis — only MediaPipe landmark extraction is performed.
    The resulting CSV is saved to data/raw/{mode}/{subject}/.
    """
    import io as _io_mod
    import tempfile

    subject_id = request.form.get("subject_id", "RAW").strip() or "RAW"
    session_label = request.form.get("session_label", "export").strip() or "export"
    study_mode = request.form.get("study_mode", "pilot").strip()

    video_path_strs = request.form.get("camera_video_paths", "").strip()
    if not video_path_strs:
        return jsonify({"status": "error", "message": "At least one camera video path is required"}), 400

    video_paths = []
    for p in video_path_strs.split("\n"):
        p = p.strip()
        if not p:
            continue
        vp = Path(p)
        if not vp.exists():
            return jsonify({"status": "error", "message": f"Video file not found: {p}"}), 400
        video_paths.append(vp)

    if len(video_paths) > 4:
        return jsonify({"status": "error", "message": "Maximum 4 camera video paths allowed"}), 400

    ts_upload_dir = UPLOAD_FOLDER / f"{subject_id}_raw_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    ts_upload_dir.mkdir(parents=True, exist_ok=True)

    timestamps_path = None
    timestamps_file = request.files.get("timestamps_csv")
    if timestamps_file and timestamps_file.filename:
        timestamps_path = ts_upload_dir / timestamps_file.filename
        timestamps_file.save(str(timestamps_path))
        if timestamps_path.suffix.lower() in (".xlsx", ".xls"):
            try:
                timestamps_path = _convert_timestamps_xlsx_to_csv(timestamps_path)
            except Exception as exc:
                return jsonify({"status": "error", "message": f"Could not convert Excel timestamps: {exc}"}), 400

    meta_path = None
    meta_file = request.files.get("recording_meta")
    if meta_file and meta_file.filename:
        meta_path = ts_upload_dir / meta_file.filename
        meta_file.save(str(meta_path))
    elif timestamps_path is not None:
        meta_path = _auto_generate_recording_meta(timestamps_path, ts_upload_dir)

    assembly_file = request.files.get("assembly_csv")
    if assembly_file and assembly_file.filename:
        assembly_save_path = ts_upload_dir / assembly_file.filename
        assembly_file.save(str(assembly_save_path))

    job_id = str(uuid.uuid4())
    q = queue.Queue()
    
    def worker():
        _run_raw_export_worker(job_id, subject_id, session_label, study_mode,
                               video_paths, timestamps_path, meta_path, q)
    
    thread = threading.Thread(target=worker, daemon=False)
    
    with _jobs_lock:
        _jobs[job_id] = {
            "queue": q,
            "thread": thread,
            "running": True,
            "summary": {},
        }
    
    thread.start()
    
    return jsonify({"status": "started", "job_id": job_id})


@app.route("/api/download_raw_csv/<path:session_path>", methods=["GET"])
def download_raw_csv(session_path: str):
    """Serve the raw frames CSV produced during a pipeline run.

    session_path is relative to data/raw/ (e.g. pilot/P001/session_id/raw_frames.csv).
    """
    raw_root = (PIPELINE_ROOT / "data" / "raw").resolve()
    requested = (raw_root / session_path).resolve()
    try:
        requested.relative_to(raw_root)
    except ValueError:
        return jsonify({"error": "Access denied"}), 403
    if not requested.exists():
        csvs = list(requested.parent.glob("*_raw_frames.csv"))
        if not csvs:
            return jsonify({"error": "Raw frames CSV not found"}), 404
        requested = csvs[0]
    return send_file(str(requested), as_attachment=True)


@app.route("/api/cross_participant", methods=["POST"])
def cross_participant_analysis():
    """Run cross-participant group analysis.

    Accepts JSON body with:
      subject_ids: list[str]  — required
      study_mode: str         — default "pilot"
      group_col: str|null     — optional column to split box plots by
    Also accepts optional 'demographics' file upload.
    """
    from src.cross_participant import compare_participants

    subject_ids_raw = request.form.get("subject_ids", "").strip()
    if not subject_ids_raw:
        try:
            body = request.get_json(silent=True) or {}
            subject_ids_raw = ",".join(body.get("subject_ids", []))
        except Exception:
            pass

    subject_ids = [s.strip() for s in subject_ids_raw.replace("\n", ",").split(",") if s.strip()]
    if not subject_ids:
        return jsonify({"status": "error", "message": "subject_ids is required"}), 400

    study_mode = request.form.get("study_mode", "pilot").strip()
    group_col = request.form.get("group_col", "").strip() or None

    demo_path: Optional[Path] = None
    demo_file = request.files.get("demographics")
    if demo_file and demo_file.filename:
        demo_upload_dir = UPLOAD_FOLDER / f"demo_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        demo_upload_dir.mkdir(parents=True, exist_ok=True)
        demo_path = demo_upload_dir / demo_file.filename
        demo_file.save(str(demo_path))

    try:
        result = compare_participants(
            project_root=PIPELINE_ROOT,
            subject_ids=subject_ids,
            study_mode=study_mode,
            demographics_path=demo_path,
            group_col=group_col,
        )
        return jsonify({
            "status": "ok",
            "n_subjects": result["n_subjects"],
            "n_subjects_with_data": result.get("n_subjects_with_data", 0),
            "summary_stats": result.get("summary_stats", {}),
            "output_paths": result.get("output_paths", {}),
        })
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500


def main() -> None:
    """Parse CLI arguments and start the Flask server with background model prefetch."""
    parser = argparse.ArgumentParser(
        description="Launch the facial analysis pipeline web UI"
    )
    parser.add_argument(
        "--port", type=int, default=5050,
        help="Port to listen on (default: 5050)",
    )
    parser.add_argument(
        "--host", default="127.0.0.1",
        help="Host to bind to (default: 127.0.0.1)",
    )
    args = parser.parse_args()

    model_thread = threading.Thread(target=_prefetch_model, daemon=True)
    model_thread.start()

    port = _find_free_port(args.port)
    if port != args.port:
        print(f"[info] Port {args.port} is in use; using {port} instead.")

    server_url = f"http://{args.host}:{port}"
    fleet_thread = threading.Thread(
        target=lambda: app.run(host=args.host, port=port, debug=False, 
                              use_reloader=False, threaded=True),
        daemon=True,
    )
    fleet_thread.start()
    print(f"[info] Flask server starting on {server_url}")

    if _wait_for_server(server_url, timeout=15.0):
        print(f"[info] Server is ready. Opening {server_url} in your browser.")
        webbrowser.open(server_url)
    else:
        print(f"[warn] Flask server did not start within 15 seconds.")
        print(f"       Open {server_url} manually in your browser.")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\n[info] Shutting down.")


if __name__ == "__main__":
    main()
